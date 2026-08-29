"""
Builds the monthly preventive-maintenance trend report from Equipment.xlsx.
"""

import calendar
import re
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from config import TEMPLATES_DIR


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
WEEKDAY_INDEX = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
    "Friday": 4, "Saturday": 5, "Sunday": 6,
}
FREQUENCY_COLUMNS = {"W": (2, 3), "M": (4, 5), "3M": (6, 7), "6M": (8, 9), "Y": (10, 11), "2Y": (12, 13)}
REQUIRED_COLUMNS = {"Equipment No.", "Maintenance Date", "Last 2Y Done", *MONTHS}


class TrendController:

    def __init__(self, excel_file, output_file, year, half, weekday, template_file=None):
        self.excel_file = Path(excel_file)
        self.output_file = Path(output_file)
        self.year = int(year)
        self.half = half
        self.weekday = weekday
        self.template_file = Path(template_file or TEMPLATES_DIR / "Template.xlsx")

    def generate(self):
        self._validate_files()
        equipment = self._load_equipment()
        months = MONTHS[:6] if self.half == "Jan-Jun" else MONTHS[6:]
        workbook = load_workbook(self.template_file)
        self._prepare_month_sheets(workbook, months)

        for month in months:
            sheet = workbook[month.upper()]
            self._fill_month_sheet(sheet, month, equipment)

        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.output_file)
        return self.output_file

    def _validate_files(self):
        if not self.excel_file.exists():
            raise FileNotFoundError(f"Equipment workbook not found:\n{self.excel_file}")
        if not self.template_file.exists():
            raise FileNotFoundError(f"Trend report template not found:\n{self.template_file}")

    def _load_equipment(self):
        source = pd.ExcelFile(self.excel_file)
        sections = {"Engineering": [], "Warehouse": [], "Production": []}
        for sheet_name in source.sheet_names:
            if sheet_name == "Settings":
                continue
            data = pd.read_excel(self.excel_file, sheet_name=sheet_name).fillna("")
            data.columns = data.columns.astype(str).str.strip()
            missing = REQUIRED_COLUMNS - set(data.columns)
            if missing:
                raise ValueError(f"Sheet '{sheet_name}' is missing: {', '.join(sorted(missing))}")
            for _, row in data.iterrows():
                equipment_id = str(row["Equipment No."]).strip()
                if equipment_id:
                    sections[self._section_for(sheet_name, equipment_id)].append(row)
        for rows in sections.values():
            rows.sort(key=lambda row: self._equipment_sort_key(row["Equipment No."]))
        return sections

    @staticmethod
    def _section_for(sheet_name, equipment_id):
        if sheet_name.startswith("PR"):
            return "Production"
        if "WH" in equipment_id.upper():
            return "Warehouse"
        return "Engineering"

    @staticmethod
    def _equipment_sort_key(equipment_id):
        return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", str(equipment_id).strip())]

    def _prepare_month_sheets(self, workbook, months):
        template = workbook.worksheets[0]
        for index, month in enumerate(months):
            sheet_name = month.upper()
            if index < len(workbook.worksheets):
                workbook.worksheets[index].title = sheet_name
            else:
                copied = workbook.copy_worksheet(template)
                copied.title = sheet_name
        while len(workbook.worksheets) > len(months):
            workbook.remove(workbook.worksheets[-1])

    def _fill_month_sheet(self, sheet, month, equipment):
        self._clear_unused_template_columns(sheet)
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))

        summary_rows = 8
        sheet.insert_rows(1, summary_rows)

        sections = (("Engineering", 1 + summary_rows), ("Warehouse", 8 + summary_rows),
                    ("Production", 15 + summary_rows))
        offset = 0
        total_rows = []
        for section_name, template_start in sections:
            start_row = template_start + offset
            rows = equipment[section_name]
            self._insert_rows_from_template(sheet, template_start, start_row, len(rows))
            total_rows.append(self._write_section(sheet, start_row, section_name, rows, month))
            offset += len(rows) - 2
        self._write_summary(sheet, month, total_rows)
        self._clear_unused_template_columns(sheet)
        sheet.freeze_panes = None
        sheet.sheet_view.showGridLines = False

    @staticmethod
    def _clear_unused_template_columns(sheet):
        """Remove empty formatting carried beyond the template's final column U."""
        for column_letter in list(sheet.column_dimensions):
            if column_index_from_string(column_letter) > 21:
                del sheet.column_dimensions[column_letter]
        for row_number, column_number in list(sheet._cells):
            if column_number > 21:
                del sheet._cells[(row_number, column_number)]

    def _insert_rows_from_template(self, sheet, template_start, start_row, equipment_count):
        needed_rows = max(equipment_count, 1)
        change = needed_rows - 2
        if change > 0:
            insert_at = start_row + 5
            sheet.insert_rows(insert_at, change)
            for offset in range(change):
                self._copy_row_style(sheet, start_row + 3, insert_at + offset)
        elif change < 0:
            sheet.delete_rows(start_row + 3 + needed_rows, -change)

    @staticmethod
    def _copy_row_style(sheet, source_row, target_row):
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
        for column in range(1, 22):
            source = sheet.cell(source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)
            if source.font:
                target.font = copy(source.font)

    def _write_section(self, sheet, start_row, section_name, rows, month):
        title_row = start_row
        header_row = start_row + 1
        data_start = start_row + 3
        total_row = data_start + max(len(rows), 1)
        sheet.cell(title_row, 1, f"{section_name} Equipment")
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=21)
        for start_column in range(2, 19, 2):
            sheet.merge_cells(start_row=header_row, start_column=start_column,
                              end_row=header_row, end_column=start_column + 1)

        for row_index, equipment in enumerate(rows, data_start):
            self._write_equipment_row(sheet, row_index, equipment, month)
        if not rows:
            self._write_empty_row(sheet, data_start)

        sheet.cell(total_row, 1, "Total")
        for column in range(2, 21):
            letter = get_column_letter(column)
            sheet.cell(total_row, column, f"=SUM({letter}{data_start}:{letter}{total_row - 1})")
        sheet.cell(total_row, 21, None)

        # Keep the original two-row headings after rows are inserted/deleted.
        sheet.cell(header_row, 1, "Equipment No.")
        return total_row

    def _write_summary(self, sheet, month, total_rows):
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill("solid", fgColor="1F4E78")
        value_fill = PatternFill("solid", fgColor="D9EAF7")
        input_fill = PatternFill("solid", fgColor="FFF2CC")

        sheet.merge_cells("A1:U1")
        title = sheet["A1"]
        title.value = f"{month.upper()} {self.year} - Preventive Maintenance Summary"
        title.fill = title_fill
        title.font = Font(bold=True, color="FFFFFF", size=13)
        title.alignment = Alignment(horizontal="center")

        labels = ["Total Scheduled", "Total Executed", "Within Tolerance", "Without Tolerance", "Total Not Executed"]
        columns = ["P", "Q", "R", "S", "T"]
        for row_number, label, column in zip(range(2, 7), labels, columns):
            sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
            label_cell = sheet.cell(row_number, 1, label)
            value_cell = sheet.cell(row_number, 5, "=" + "+".join(f"{column}{row}" for row in total_rows))
            for cell in (label_cell, value_cell):
                cell.border = border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            value_cell.fill = value_fill

        inputs = [("F2:H2", "Deviation"), ("F3:H3", "Production Plant"),
                  ("J2:L2", "Long Time No Usage"), ("J3:L3", "CCR")]
        for cell_range, label in inputs:
            sheet.merge_cells(cell_range)
            label_cell = sheet[cell_range.split(":")[0]]
            label_cell.value = label
            label_cell.border = border
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="center")
            start, end = cell_range.split(":")
            value_column = get_column_letter(column_index_from_string(end[0]) + 1)
            value_cell = sheet[f"{value_column}{start[1:]}"]
            value_cell.value = None
            value_cell.fill = input_fill
            value_cell.border = border

    def _write_equipment_row(self, sheet, row_number, equipment, month):
        sheet.cell(row_number, 1, str(equipment["Equipment No."]).strip())
        for column in range(2, 21):
            sheet.cell(row_number, column, 0)
        sheet.cell(row_number, 21, None)

        tokens = self._tokens(equipment[month])
        for frequency, (scheduled_column, executed_column) in FREQUENCY_COLUMNS.items():
            value = self._scheduled_count(frequency, month, tokens, equipment["Last 2Y Done"])
            sheet.cell(row_number, scheduled_column, value)
            sheet.cell(row_number, executed_column, value)

        sheet.cell(row_number, 16, f"=SUM(B{row_number},D{row_number},F{row_number},H{row_number},J{row_number},L{row_number},N{row_number})")
        sheet.cell(row_number, 17, f"=SUM(C{row_number},E{row_number},G{row_number},I{row_number},K{row_number},M{row_number},O{row_number})")
        sheet.cell(row_number, 18, f"=Q{row_number}")
        sheet.cell(row_number, 19, 0)
        sheet.cell(row_number, 20, f"=P{row_number}-Q{row_number}")

    @staticmethod
    def _write_empty_row(sheet, row_number):
        for column in range(1, 22):
            sheet.cell(row_number, column, 0 if column > 1 and column < 21 else None)

    @staticmethod
    def _tokens(value):
        return {token.strip().upper() for token in str(value).split(",") if token.strip() and token.strip() != "-"}

    def _scheduled_count(self, frequency, month, tokens, last_two_year_done):
        if frequency not in tokens:
            return 0
        if frequency == "W":
            month_number = MONTHS.index(month) + 1
            # Trend reports always use Friday for weekly maintenance.
            friday_index = WEEKDAY_INDEX["Friday"]
            return sum(
                day == friday_index
                for week in calendar.monthcalendar(self.year, month_number)
                for day in week
                if day
            )
        if frequency == "2Y" and not self._two_yearly_due(last_two_year_done):
            return 0
        return 1

    def _two_yearly_due(self, value):
        if value in (None, "") or pd.isna(value):
            return True
        try:
            year = int(float(value))
        except (TypeError, ValueError):
            parsed = pd.to_datetime(value, errors="coerce")
            if pd.isna(parsed):
                return True
            year = parsed.year
        return self.year - year >= 2
